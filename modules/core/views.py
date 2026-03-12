"""
Core views - Dashboard, Login, Logout, Reports
"""
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt
from django.views.decorators.http import require_http_methods
from inertia import inertia, render as inertia_render

@login_required
@inertia('Index')
def index(request):
    """Dashboard with ticket performance metrics."""
    from django.db.models import Count, Avg, Q, F
    from django.db.models.functions import TruncMonth, TruncDate, TruncWeek
    from django.utils import timezone
    from datetime import timedelta
    from modules.ticket.models import Ticket, ActivityStream
    from modules.users.models import Group
    
    user_profile = None
    try:
        user_profile = request.user.profile
    except:
        pass

    # Get time filter from query params - default to 7 days
    time_filter = request.GET.get('timeFilter', '7days')
    
    now = timezone.now()
    
    # Calculate date range based on filter
    filter_days = {
        '7days': 7,
        '30days': 30,
        '3months': 90,
        '6months': 180,
        '1year': 365,
    }
    days = filter_days.get(time_filter, 7)
    start_date = now - timedelta(days=days)
    
    # === TICKET STATS (filtered by time) ===
    all_tickets = Ticket.objects.all()
    filtered_tickets = all_tickets.filter(created_at__gte=start_date)
    
    # Count by status - map to frontend expected keys
    new_count = filtered_tickets.filter(status='new').count()
    open_count = filtered_tickets.filter(status='open').count()
    in_progress_count = filtered_tickets.filter(status='in_progress').count()
    pending_count = filtered_tickets.filter(status='pending').count()
    resolved_count = filtered_tickets.filter(status='resolved').count()
    closed_count = filtered_tickets.filter(status='closed').count()
    total_tickets = filtered_tickets.count()
    
    # Status distribution with percentages
    status_distribution = []
    if total_tickets > 0:
        for status, label in [('new', 'New'), ('open', 'Open'), ('in_progress', 'In Progress'), 
                               ('pending', 'On Hold'), ('resolved', 'Resolved'), ('closed', 'Closed')]:
            count = filtered_tickets.filter(status=status).count()
            if count > 0:
                status_distribution.append({
                    'status': label,
                    'count': count,
                    'percentage': round((count / total_tickets) * 100, 1)
                })
    
    # Priority breakdown
    priority_counts = filtered_tickets.exclude(status__in=['resolved', 'closed']).values('priority').annotate(count=Count('id'))
    open_total = sum(p['count'] for p in priority_counts)
    priorities = []
    for p in priority_counts:
        if p['count'] > 0:
            priorities.append({
                'label': p['priority'].capitalize() if p['priority'] else 'Normal',
                'count': p['count'],
                'percentage': round((p['count'] / open_total) * 100, 1) if open_total > 0 else 0
            })
    # Sort priorities by urgency
    priority_order = {'urgent': 0, 'high': 1, 'normal': 2, 'low': 3}
    priorities.sort(key=lambda x: priority_order.get(x['label'].lower(), 99))
    
    # === TIME-BASED PERFORMANCE DATA ===
    # Determine grouping based on filter
    performance_data = []
    
    if time_filter in ['7days', '30days']:
        # Daily grouping
        daily_data = filtered_tickets.annotate(
            day=TruncDate('created_at')
        ).values('day').annotate(
            total=Count('id'),
        ).order_by('day')
        
        # Create a complete date range
        date_labels = []
        current = start_date.date()
        end = now.date()
        while current <= end:
            date_labels.append(current)
            current += timedelta(days=1)
        
        # Map data to dates
        daily_dict = {}
        for d in daily_data:
            daily_dict[d['day']] = d['total']
        
        for date in date_labels:
            day_tickets = filtered_tickets.filter(created_at__date=date)
            performance_data.append({
                'label': date.strftime('%b %d') if time_filter == '30days' else date.strftime('%a'),
                'fullDate': date.strftime('%Y-%m-%d'),
                'open': day_tickets.filter(status__in=['new', 'open']).count(),
                'inProgress': day_tickets.filter(status='in_progress').count(),
                'hold': day_tickets.filter(status='pending').count(),
                'resolved': day_tickets.filter(status='resolved').count(),
                'closed': day_tickets.filter(status='closed').count(),
            })
    
    elif time_filter in ['3months', '6months']:
        # Weekly grouping
        weekly_data = filtered_tickets.annotate(
            week=TruncWeek('created_at')
        ).values('week').annotate(
            total=Count('id'),
        ).order_by('week')
        
        # Build complete week range
        weeks_count = days // 7
        for i in range(weeks_count):
            week_start = start_date + timedelta(weeks=i)
            week_end = week_start + timedelta(days=7)
            week_tickets = filtered_tickets.filter(
                created_at__gte=week_start,
                created_at__lt=week_end
            )
            performance_data.append({
                'label': f"W{i+1}",
                'fullDate': week_start.strftime('%b %d'),
                'open': week_tickets.filter(status__in=['new', 'open']).count(),
                'inProgress': week_tickets.filter(status='in_progress').count(),
                'hold': week_tickets.filter(status='pending').count(),
                'resolved': week_tickets.filter(status='resolved').count(),
                'closed': week_tickets.filter(status='closed').count(),
            })
    
    else:  # 1year - monthly grouping
        monthly_data = filtered_tickets.annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=Count('id'),
        ).order_by('month')
        
        for m in monthly_data:
            month_start = m['month']
            month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            month_tickets = filtered_tickets.filter(created_at__gte=month_start, created_at__lt=month_end)
            performance_data.append({
                'label': month_start.strftime('%b'),
                'fullDate': month_start.strftime('%Y-%m'),
                'open': month_tickets.filter(status__in=['new', 'open']).count(),
                'inProgress': month_tickets.filter(status='in_progress').count(),
                'hold': month_tickets.filter(status='pending').count(),
                'resolved': month_tickets.filter(status='resolved').count(),
                'closed': month_tickets.filter(status='closed').count(),
            })
    
    # === SLA STATS ===
    # Calculate average resolution time from resolved tickets
    resolved_tickets = all_tickets.filter(
        status__in=['resolved', 'closed'],
        resolved_at__isnull=False
    )
    
    avg_resolution_time = "—"
    if resolved_tickets.exists():
        avg_resolution_seconds = resolved_tickets.aggregate(
            avg=Avg(F('resolved_at') - F('created_at'))
        )['avg']
        if avg_resolution_seconds:
            total_seconds = avg_resolution_seconds.total_seconds()
            hours = int(total_seconds // 3600)
            minutes = int((total_seconds % 3600) // 60)
            avg_resolution_time = f"{hours}h {minutes}m"
    
    # Count open tickets for SLA stats (simplified - show active tickets status)
    active_tickets = all_tickets.exclude(status__in=['resolved', 'closed']).count()
    
    sla_data = {
        'breaches': 0,  # Would need SLA tracking fields on Ticket model
        'onTrack': active_tickets,
        'atRisk': 0,  # Would need SLA tracking fields on Ticket model
        'avgResponseTime': "—",  # Would need first_response_at field
        'avgResolutionTime': avg_resolution_time,
        'complianceRate': 100,  # Would need SLA tracking fields
    }
    
    # === GROUP STATS ===
    groups = Group.objects.annotate(
        active_tickets=Count('ticket', filter=Q(ticket__status__in=['new', 'open', 'in_progress', 'pending']))
    ).filter(active_tickets__gt=0).order_by('-active_tickets')[:5]
    
    group_stats = []
    for group in groups:
        # Calculate avg resolution time for this group
        group_resolved = all_tickets.filter(
            group=group,
            status__in=['resolved', 'closed'],
            resolved_at__isnull=False
        )
        if group_resolved.exists():
            avg_time = group_resolved.aggregate(
                avg=Avg(F('resolved_at') - F('created_at'))
            )['avg']
            if avg_time:
                total_seconds = avg_time.total_seconds()
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                avg_res_time = f"{hours}h {minutes}m"
            else:
                avg_res_time = "—"
        else:
            avg_res_time = "—"
        
        group_stats.append({
            'name': group.name,
            'activeTickets': group.active_tickets,
            'avgResolutionTime': avg_res_time,
        })
    
    # === RECENT ACTIVITY ===
    recent_activities = ActivityStream.objects.select_related(
        'actor', 'ticket'
    ).order_by('-created_at')[:5]
    
    recent_activity = []
    for activity in recent_activities:
        action = activity.description or f"Activity on ticket"
        if activity.ticket:
            action = f"{action}"
        
        user_name = "System"
        if activity.actor:
            user_name = f"{activity.actor.first_name} {activity.actor.last_name}".strip() or activity.actor.username
        
        # Format time ago
        time_diff = now - activity.created_at
        if time_diff.days > 0:
            time_ago = f"{time_diff.days} day{'s' if time_diff.days > 1 else ''} ago"
        elif time_diff.seconds >= 3600:
            hours = time_diff.seconds // 3600
            time_ago = f"{hours} hour{'s' if hours > 1 else ''} ago"
        elif time_diff.seconds >= 60:
            minutes = time_diff.seconds // 60
            time_ago = f"{minutes} minute{'s' if minutes > 1 else ''} ago"
        else:
            time_ago = "Just now"
        
        recent_activity.append({
            'action': action,
            'user': user_name,
            'time': time_ago,
        })

    return {
        'title': 'Dashboard',
        'message': 'Welcome to ImaraDesk!',
        'timeFilter': time_filter,
        'dashboard': {
            'tickets': {
                'open': new_count + open_count,  # Frontend shows "Open" as new + open
                'inProgress': in_progress_count,
                'hold': pending_count,
                'resolved': resolved_count,
                'closed': closed_count,
                'total': total_tickets,
            },
            'sla': sla_data,
            'performanceData': performance_data,
            'statusDistribution': status_distribution,
            'priorities': priorities,
            'groups': group_stats,
            'recentActivity': recent_activity,
        },
    }

@ensure_csrf_cookie
@inertia('Login')
def login_view(request):
    """User login page (GET + POST)."""
    # Check if organization is verified before allowing login attempts
    try:
        from shared.models import Client
        tenant = Client.get_current()
        if tenant and hasattr(tenant, 'is_verified') and not tenant.is_verified:
            return inertia_render(request, 'Login', {
                'errors': {'general': 'This workspace has not been verified yet. Please check your email for the verification link.'},
            })
    except Exception:
        pass  # If we can't check, proceed with normal login
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        errors = {}

        if not email:
            errors['email'] = 'Email is required'
        if not password:
            errors['password'] = 'Password is required'

        if errors:
            return inertia_render(request, 'Login', {
                'errors': errors,
            })

        # Try to find user by email
        try:
            user_obj = User.objects.get(email=email)
            username = user_obj.username
        except User.DoesNotExist:
            return inertia_render(request, 'Login', {
                'errors': {'general': 'Invalid email or password'},
            })

        # Authenticate
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Check if user has 2FA enabled
            try:
                profile = user.profile
                email_2fa = profile.email_2fa_enabled
                auth_2fa = profile.authenticator_2fa_enabled
                
                if email_2fa or auth_2fa:
                    # Store user ID and 2FA methods in session for 2FA verification
                    request.session['pending_2fa_user_id'] = user.id
                    request.session['pending_2fa_email'] = email_2fa
                    request.session['pending_2fa_authenticator'] = auth_2fa
                    request.session['pending_2fa_next'] = request.POST.get('next') or request.GET.get('next') or 'index'
                    
                    # Redirect to 2FA verification page
                    return redirect('two_factor_verify')
            except:
                pass
            
            # No 2FA - proceed with normal login
            login(request, user)

            # Handle redirect: check for 'next' parameter
            next_url = request.POST.get('next') or request.GET.get('next') or 'index'

            # Security check: ensure the URL is safe
            if next_url.startswith('/'):
                return redirect(next_url)
            else:
                return redirect('index')
        else:
            return inertia_render(request, 'Login', {
                'errors': {'general': 'Invalid email or password'},
            })

    if request.user.is_authenticated:
        return redirect('index')
    # GET request - pass next parameter to the form
    next_url = request.GET.get('next', '')
    return {
        'errors': {},
        'next': next_url,
    }


def logout_view(request):
    """Logout the user and redirect to login page."""
    logout(request)
    return redirect('login')


@ensure_csrf_cookie
@inertia('TwoFactorVerify')
def two_factor_verify(request):
    """2FA verification page after successful password authentication."""
    # Check if there's a pending 2FA verification
    user_id = request.session.get('pending_2fa_user_id')
    if not user_id:
        return redirect('login')
    
    try:
        user = User.objects.get(id=user_id)
        profile = user.profile
    except (User.DoesNotExist, Exception):
        # Clear session and redirect to login
        request.session.pop('pending_2fa_user_id', None)
        return redirect('login')
    
    email_2fa = request.session.get('pending_2fa_email', False)
    auth_2fa = request.session.get('pending_2fa_authenticator', False)
    
    return {
        'email_2fa_enabled': email_2fa,
        'authenticator_2fa_enabled': auth_2fa,
        'user_email': user.email,
        'user_name': profile.full_name or user.username,
    }


@csrf_exempt
@require_http_methods(["POST"])
def send_login_2fa_code(request):
    """Send 2FA code for login verification."""
    import json
    import random
    from django.utils import timezone
    from datetime import timedelta
    from shared.utilities.Mailer import Mailer
    from shared.utilities.GlobalEmailTenplates import GlobalEmailTemplates
    
    user_id = request.session.get('pending_2fa_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'message': 'No pending login'}, status=400)
    
    try:
        user = User.objects.get(id=user_id)
        profile = user.profile
        
        # Generate 6-digit code
        code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        
        # Store code with expiration (10 minutes)
        profile.email_2fa_code = code
        profile.email_2fa_code_expires = timezone.now() + timedelta(minutes=10)
        profile.save()
        
        # Send email using Mailer with GlobalEmailTemplates
        template = GlobalEmailTemplates.TWO_FA_LOGIN_CODE
        mailer = Mailer()
        mailer.send_raw_email(
            to_email=user.email,
            subject=template['subject'],
            body_html=template['body_html'],
            body_text=template['body_text'],
            context={
                'user_name': profile.full_name or user.username,
                'code': code,
            },
            fail_silently=False,
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Verification code sent to your email'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': 'Failed to send verification code'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def verify_login_2fa(request):
    """Verify 2FA code and complete login."""
    import json
    from django.utils import timezone
    import pyotp
    
    user_id = request.session.get('pending_2fa_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'message': 'No pending login'}, status=400)
    
    try:
        data = json.loads(request.body)
        code = data.get('code', '')
        method = data.get('method', 'email')  # 'email' or 'authenticator'
        
        user = User.objects.get(id=user_id)
        profile = user.profile
        
        if method == 'email':
            # Verify email code
            if not profile.email_2fa_code:
                return JsonResponse({
                    'success': False,
                    'message': 'No verification code found. Please request a new one.'
                }, status=400)
            
            if profile.email_2fa_code_expires and profile.email_2fa_code_expires < timezone.now():
                return JsonResponse({
                    'success': False,
                    'message': 'Verification code has expired. Please request a new one.'
                }, status=400)
            
            if profile.email_2fa_code != code:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid verification code'
                }, status=400)
            
            # Clear the code
            profile.email_2fa_code = ''
            profile.email_2fa_code_expires = None
            profile.save()
            
        else:  # authenticator
            # Verify TOTP code
            if not profile.totp_secret:
                return JsonResponse({
                    'success': False,
                    'message': 'Authenticator not configured'
                }, status=400)
            
            totp = pyotp.TOTP(profile.totp_secret)
            if not totp.verify(code, valid_window=1):
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid verification code'
                }, status=400)
        
        # Code verified - complete login
        login(request, user)
        
        # Get redirect URL and clean up session
        next_url = request.session.get('pending_2fa_next', 'index')
        request.session.pop('pending_2fa_user_id', None)
        request.session.pop('pending_2fa_email', None)
        request.session.pop('pending_2fa_authenticator', None)
        request.session.pop('pending_2fa_next', None)
        
        return JsonResponse({
            'success': True,
            'redirect': next_url if next_url.startswith('/') else f'/{next_url}/' if next_url != 'index' else '/dashboard/'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'User not found'}, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': 'Verification failed'
        }, status=500)


@login_required
@inertia('Reports')
def reports(request):
    """Reports and analytics page with comprehensive data."""
    from django.utils import timezone
    from django.db.models import Count, Avg, Q, F
    from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
    from datetime import timedelta, datetime
    from modules.ticket.models import Ticket
    from modules.users.models import Group
    
    # Get report type and date filters from query params
    report_type = request.GET.get('type', 'tickets')
    date_range = request.GET.get('range', '30')  # days (preset)
    
    # Custom date range support
    custom_start = request.GET.get('start_date')
    custom_end = request.GET.get('end_date')
    
    # Filter params
    group_filter = request.GET.get('group')
    agent_filter = request.GET.get('agent')
    status_filter = request.GET.get('status')
    priority_filter = request.GET.get('priority')
    
    now = timezone.now()
    
    # Determine date range
    if custom_start and custom_end:
        try:
            start_date = timezone.make_aware(datetime.fromisoformat(custom_start.replace('Z', '')))
            end_date = timezone.make_aware(datetime.fromisoformat(custom_end.replace('Z', '')))
            date_range = 'custom'
        except:
            days = 30
            start_date = now - timedelta(days=days)
            end_date = now
    else:
        try:
            days = int(date_range)
        except:
            days = 30
        start_date = now - timedelta(days=days)
        end_date = now
    
    # Get filter options (groups and agents)
    groups_list = list(Group.objects.values('id', 'name').order_by('name'))
    agents_list = list(User.objects.filter(
        is_active=True, 
        profile__is_agent=True
    ).values('id', 'first_name', 'last_name', 'username').order_by('first_name', 'last_name'))
    
    # Format agent names
    for agent in agents_list:
        name = f"{agent['first_name']} {agent['last_name']}".strip()
        agent['name'] = name if name else agent['username']
    
    # Initialize response data
    data = {
        'report_type': report_type,
        'date_range': date_range,
        'filters': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'group': group_filter,
            'agent': agent_filter,
            'status': status_filter,
            'priority': priority_filter,
        },
        'filter_options': {
            'groups': groups_list,
            'agents': agents_list,
            'statuses': [
                {'id': 'new', 'name': 'New'},
                {'id': 'open', 'name': 'Open'},
                {'id': 'in_progress', 'name': 'In Progress'},
                {'id': 'pending', 'name': 'Pending'},
                {'id': 'resolved', 'name': 'Resolved'},
                {'id': 'closed', 'name': 'Closed'},
            ],
            'priorities': [
                {'id': 'low', 'name': 'Low'},
                {'id': 'normal', 'name': 'Normal'},
                {'id': 'high', 'name': 'High'},
                {'id': 'urgent', 'name': 'Urgent'},
            ],
        },
    }
    
    # ========== TICKET REPORTS ==========
    if report_type == 'tickets':
        all_tickets = Ticket.objects.all()
        
        # Apply filters
        if group_filter:
            all_tickets = all_tickets.filter(group_id=group_filter)
        if agent_filter:
            all_tickets = all_tickets.filter(assignee_id=agent_filter)
        if status_filter:
            all_tickets = all_tickets.filter(status=status_filter)
        if priority_filter:
            all_tickets = all_tickets.filter(priority=priority_filter)
        
        period_tickets = all_tickets.filter(created_at__gte=start_date, created_at__lte=end_date)
        
        # Basic metrics
        total = all_tickets.count()
        period_total = period_tickets.count()
        
        # Status breakdown
        status_counts = all_tickets.values('status').annotate(count=Count('id')).order_by('-count')
        
        # Priority breakdown
        priority_counts = all_tickets.values('priority').annotate(count=Count('id')).order_by('-count')
        
        # Source breakdown
        source_counts = all_tickets.values('source').annotate(count=Count('id')).order_by('-count')
        
        # Type breakdown
        type_counts = all_tickets.values('type').annotate(count=Count('id')).order_by('-count')
        
        # Aging tickets (open tickets by age)
        aging_buckets = {
            '0-1 day': 0,
            '1-3 days': 0,
            '3-7 days': 0,
            '7-14 days': 0,
            '14-30 days': 0,
            '30+ days': 0,
        }
        
        open_statuses = ['new', 'open', 'in_progress', 'pending']
        open_tickets = all_tickets.filter(status__in=open_statuses)
        
        for ticket in open_tickets:
            age = (now - ticket.created_at).days
            if age < 1:
                aging_buckets['0-1 day'] += 1
            elif age < 3:
                aging_buckets['1-3 days'] += 1
            elif age < 7:
                aging_buckets['3-7 days'] += 1
            elif age < 14:
                aging_buckets['7-14 days'] += 1
            elif age < 30:
                aging_buckets['14-30 days'] += 1
            else:
                aging_buckets['30+ days'] += 1
        
        # SLA metrics
        sla_data = {
            'total_with_sla': 0,
            'response_breached': 0,
            'resolution_breached': 0,
            'on_track': 0,
            'compliance_rate': 0,
        }
        
        tickets_with_sla = all_tickets.filter(sla__isnull=False).select_related('sla')
        sla_data['total_with_sla'] = tickets_with_sla.count()
        
        if sla_data['total_with_sla'] > 0:
            sla_data['response_breached'] = tickets_with_sla.filter(sla__response_breached=True).count()
            sla_data['resolution_breached'] = tickets_with_sla.filter(sla__resolution_breached=True).count()
            sla_data['on_track'] = tickets_with_sla.filter(
                sla__response_breached=False, 
                sla__resolution_breached=False
            ).count()
            sla_data['compliance_rate'] = round(
                (sla_data['on_track'] / sla_data['total_with_sla']) * 100, 1
            )
        
        # Daily ticket volume for trend chart
        daily_volume = period_tickets.annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(count=Count('id')).order_by('date')
        
        # Resolution metrics
        resolved_tickets = all_tickets.filter(status__in=['resolved', 'closed'], resolved_at__isnull=False)
        
        # Group performance
        group_stats = all_tickets.filter(group__isnull=False).values(
            'group__name'
        ).annotate(
            total=Count('id'),
            open=Count('id', filter=Q(status__in=open_statuses)),
        ).order_by('-total')[:10]
        
        # Assignee performance
        assignee_stats = all_tickets.filter(assignee__isnull=False).values(
            'assignee__first_name', 'assignee__last_name', 'assignee__username'
        ).annotate(
            total=Count('id'),
            resolved=Count('id', filter=Q(status__in=['resolved', 'closed'])),
            open=Count('id', filter=Q(status__in=open_statuses)),
        ).order_by('-total')[:10]
        
        # Unassigned tickets count
        unassigned_count = all_tickets.filter(assignee__isnull=True).count()
        
        # On hold/pending tickets count
        on_hold_count = all_tickets.filter(status='pending').count()
        
        # Unassigned tickets list for report
        unassigned_tickets = list(all_tickets.filter(assignee__isnull=True).values(
            'id', 'ticket_number', 'title', 'priority', 'status', 'created_at', 
            'requester__first_name', 'requester__last_name', 'requester__email'
        ).order_by('-created_at')[:50])
        
        # On hold tickets list for report
        on_hold_tickets = list(all_tickets.filter(status='pending').values(
            'id', 'ticket_number', 'title', 'priority', 'created_at',
            'requester__first_name', 'requester__last_name', 'requester__email',
            'assignee__first_name', 'assignee__last_name'
        ).order_by('-created_at')[:50])
        
        data['tickets'] = {
            'total': total,
            'period_total': period_total,
            'open_count': open_tickets.count(),
            'unassigned_count': unassigned_count,
            'on_hold_count': on_hold_count,
            'unassigned_tickets': unassigned_tickets,
            'on_hold_tickets': on_hold_tickets,
            'status_breakdown': list(status_counts),
            'priority_breakdown': list(priority_counts),
            'source_breakdown': list(source_counts),
            'type_breakdown': list(type_counts),
            'aging': aging_buckets,
            'sla': sla_data,
            'daily_volume': list(daily_volume),
            'group_stats': list(group_stats),
            'assignee_stats': [{
                'name': f"{a['assignee__first_name']} {a['assignee__last_name']}".strip() or a['assignee__username'],
                'total': a['total'],
                'resolved': a['resolved'],
                'open': a['open'],
            } for a in assignee_stats],
        }
    
    return data


@login_required
def export_reports(request):
    """Export report data to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Count, Q, Avg
    from datetime import timedelta
    from django.utils import timezone
    
    report_type = request.GET.get('type', 'tickets')
    date_range = request.GET.get('range', '30')
    
    # Parse date range
    now = timezone.now()
    if date_range == 'custom':
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc) if start_date_str else now - timedelta(days=30)
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=timezone.utc) if end_date_str else now
        except:
            start_date = now - timedelta(days=30)
            end_date = now
    elif date_range == '7':
        start_date = now - timedelta(days=7)
        end_date = now
    elif date_range == '90':
        start_date = now - timedelta(days=90)
        end_date = now
    elif date_range == '365':
        start_date = now - timedelta(days=365)
        end_date = now
    else:
        start_date = now - timedelta(days=30)
        end_date = now
    
    # Get filters
    group_filter = request.GET.get('group', '')
    agent_filter = request.GET.get('agent', '')
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4a154b', end_color='4a154b', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header_row(worksheet, num_cols):
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def auto_adjust_columns(worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    if report_type == 'tickets':
        from modules.ticket.models import Ticket
        
        ws.title = 'Tickets Report'
        
        # Build queryset with filters
        tickets = Ticket.objects.filter(created_at__gte=start_date, created_at__lte=end_date)
        
        if group_filter:
            tickets = tickets.filter(group_id=group_filter)
        if agent_filter:
            tickets = tickets.filter(assignee_id=agent_filter)
        if status_filter:
            tickets = tickets.filter(status=status_filter)
        if priority_filter:
            tickets = tickets.filter(priority=priority_filter)
        
        tickets = tickets.select_related('requester', 'assignee', 'group').order_by('-created_at')
        
        # Headers
        headers = ['Ticket #', 'Title', 'Status', 'Priority', 'Type', 'Requester', 'Requester Email', 'Assignee', 'Group', 'Created', 'Updated']
        ws.append(headers)
        style_header_row(ws, len(headers))
        
        # Data
        for ticket in tickets:
            ws.append([
                ticket.ticket_number,
                ticket.title,
                ticket.get_status_display() if hasattr(ticket, 'get_status_display') else ticket.status,
                ticket.get_priority_display() if hasattr(ticket, 'get_priority_display') else ticket.priority,
                ticket.get_type_display() if hasattr(ticket, 'get_type_display') else (ticket.type or ''),
                f"{ticket.requester.first_name} {ticket.requester.last_name}".strip() if ticket.requester else '',
                ticket.requester.email if ticket.requester else '',
                f"{ticket.assignee.first_name} {ticket.assignee.last_name}".strip() if ticket.assignee else 'Unassigned',
                ticket.group.name if ticket.group else '',
                ticket.created_at.strftime('%Y-%m-%d %H:%M') if ticket.created_at else '',
                ticket.updated_at.strftime('%Y-%m-%d %H:%M') if ticket.updated_at else '',
            ])
        
        # Add summary sheet
        ws_summary = wb.create_sheet('Summary')
        ws_summary.append(['Metric', 'Value'])
        style_header_row(ws_summary, 2)
        
        total = tickets.count()
        status_counts = Ticket.objects.filter(created_at__gte=start_date, created_at__lte=end_date)
        if group_filter:
            status_counts = status_counts.filter(group_id=group_filter)
        if agent_filter:
            status_counts = status_counts.filter(assignee_id=agent_filter)
        if status_filter:
            status_counts = status_counts.filter(status=status_filter)
        if priority_filter:
            status_counts = status_counts.filter(priority=priority_filter)
        
        ws_summary.append(['Total Tickets', total])
        ws_summary.append(['Date Range', f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"])
        ws_summary.append([''])
        ws_summary.append(['Applied Filters', ''])
        if group_filter:
            ws_summary.append(['Group', group_filter])
        if agent_filter:
            ws_summary.append(['Agent', agent_filter])
        if status_filter:
            ws_summary.append(['Status', status_filter])
        if priority_filter:
            ws_summary.append(['Priority', priority_filter])
        
        auto_adjust_columns(ws_summary)
    
    auto_adjust_columns(ws)
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{report_type}_report_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@inertia('AI')
def ai(request):
    """AI assistant page."""
    return {
        'conversation': [],
    }


@inertia('Onboarding')
def onboarding(request):
    """
    Onboarding page for first-time setup.
    Shows if no business is registered in the system.
    """
    from shared.models import Client
    from django.core.management import call_command
    from modules.users.models import Group, UserProfile
    
    # If business already exists, redirect to login
    if Client.objects.exists():
        return redirect('login')
    
    if request.method == 'POST':
        # Extract form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        workspace_name = request.POST.get('workspace_name', '').strip()
        business_type = request.POST.get('business_type', '')
        org_size = request.POST.get('org_size', '')
        
        # Validation
        errors = {}
        
        if not first_name:
            errors['first_name'] = 'First name is required'
        if not last_name:
            errors['last_name'] = 'Last name is required'
        if not email:
            errors['email'] = 'Email is required'
        if not password:
            errors['password'] = 'Password is required'
        elif len(password) < 8:
            errors['password'] = 'Password must be at least 8 characters'
        if not workspace_name:
            errors['workspace_name'] = 'Workspace name is required'
        
        # Check if email already exists
        if email and User.objects.filter(email=email).exists():
            errors['email'] = 'A user with this email already exists'
        
        if errors:
            return inertia_render(request, 'Onboarding', {'errors': errors})
        
        try:
            # Create the organization/client
            client = Client.objects.create(
                name=workspace_name,
                description=f"Workspace for {workspace_name}",
                is_active=True,
                is_verified=True,
                business_type=business_type,
                org_size=org_size,
                created_by_email=email,
                created_by_name=f"{first_name} {last_name}",
            )
            print(f"✓ Created organization: {workspace_name}")
            
            # Create admin user
            username = email.split('@')[0]
            # Ensure unique username
            base_username = username
            counter = 1
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{counter}"
                counter += 1
            
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_staff=True,
                is_superuser=True,
            )
            print(f"✓ Created admin user: {email}")
            
            # Create user profile
            try:
                # Create or get Administrator group
                admin_group, _ = Group.objects.get_or_create(name='Administrator')
                
                # Create user profile
                profile = UserProfile.objects.create(
                    user=user,
                    group=admin_group,
                    phone='',
                    language='en',
                    timezone='UTC',
                )
                print(f"✓ Created user profile for: {email}")
            except Exception as e:
                print(f"Warning: Could not create user profile: {e}")
            
            # Run coredesk --init to seed initial data
            print("\n--- Running initial setup (coredesk --init) ---")
            try:
                call_command('coredesk', '--init')
                print("✓ Initial data seeded successfully")
            except Exception as e:
                print(f"Warning: Some seed data may have failed: {e}")
            
            print("\n" + "="*60)
            print("ONBOARDING COMPLETE")
            print("="*60)
            print(f"Organization: {workspace_name}")
            print(f"Admin Email:  {email}")
            print("="*60 + "\n")
            
            # Log the user in automatically
            login(request, user)
            
            # Redirect to dashboard
            return redirect('index')
            
        except Exception as e:
            print(f"✗ Onboarding error: {e}")
            import traceback
            traceback.print_exc()
            errors['general'] = f'Setup failed: {str(e)}'
            return inertia_render(request, 'Onboarding', {'errors': errors})
    
    # GET request - show the onboarding form
    return {
        'errors': {},
    }


# Error handlers
def page_not_found(request, exception):
    """Custom 404 handler."""
    return render(request, '404.html', status=404)


def server_error(request):
    """Custom 500 handler."""
    return render(request, '500.html', status=500)

